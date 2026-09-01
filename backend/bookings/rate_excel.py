from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from bookings.models import ClientProfessionRate, SiteProfessionRate


CLIENT_RATE_HEADERS = (
    "rate_id",
    "client_id",
    "client_name",
    "profession_id",
    "profession_name",
    "pay_rate",
    "bill_rate",
)
SITE_RATE_HEADERS = (
    "rate_id",
    "site_id",
    "client_name",
    "site_name",
    "profession_id",
    "profession_name",
    "pay_rate",
    "bill_rate",
)
MAX_IMPORT_ROWS = 5000
MAX_RATE = Decimal("99999999.99")
CENT = Decimal("0.01")


class RateWorkbookError(Exception):
    def __init__(self, errors):
        self.errors = list(errors)
        super().__init__("; ".join(self.errors))


@dataclass
class RateImportSummary:
    created: int
    updated: int
    entries: list


def _style_sheet(sheet, widths):
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _add_table(sheet, name):
    if sheet.max_row < 2:
        return
    table = Table(displayName=name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def build_client_rate_workbook(rate_queryset, client_queryset, profession_queryset):
    workbook = Workbook()
    rates = workbook.active
    rates.title = "Client Rates"
    rates.append(CLIENT_RATE_HEADERS)
    for rate in rate_queryset.select_related("client", "profession").order_by(
        "client__name", "profession__name", "id"
    ):
        rates.append([
            rate.id,
            rate.client_id,
            rate.client.name,
            rate.profession_id,
            rate.profession.name,
            rate.pay_rate,
            rate.bill_rate,
        ])
    _style_sheet(
        rates,
        {"A": 12, "B": 12, "C": 34, "D": 15, "E": 28, "F": 15, "G": 15},
    )
    _add_table(rates, "ClientRates")

    clients = workbook.create_sheet("Clients")
    clients.append(("client_id", "client_name", "city", "region"))
    for client in client_queryset.order_by("name", "id"):
        clients.append((client.id, client.name, client.city, client.region))
    _style_sheet(clients, {"A": 12, "B": 34, "C": 24, "D": 24})
    _add_table(clients, "ClientReferences")

    professions = workbook.create_sheet("Professions")
    professions.append(("profession_id", "profession_name"))
    for profession in profession_queryset.order_by("name", "id"):
        professions.append((profession.id, profession.name))
    _style_sheet(professions, {"A": 15, "B": 30})
    _add_table(professions, "ProfessionReferences")
    return _workbook_bytes(workbook)


def build_site_rate_workbook(rate_queryset, site_queryset, profession_queryset):
    workbook = Workbook()
    rates = workbook.active
    rates.title = "Facility Rates"
    rates.append(SITE_RATE_HEADERS)
    for rate in rate_queryset.select_related(
        "site__client", "profession"
    ).order_by("site__client__name", "site__name", "profession__name", "id"):
        rates.append([
            rate.id,
            rate.site_id,
            rate.site.client.name,
            rate.site.name,
            rate.profession_id,
            rate.profession.name,
            rate.pay_rate,
            rate.bill_rate,
        ])
    _style_sheet(
        rates,
        {
            "A": 12,
            "B": 12,
            "C": 32,
            "D": 28,
            "E": 15,
            "F": 28,
            "G": 15,
            "H": 15,
        },
    )
    _add_table(rates, "FacilityRates")

    sites = workbook.create_sheet("Facilities")
    sites.append(("site_id", "client_name", "site_name"))
    for site in site_queryset.select_related("client").order_by(
        "client__name", "name", "id"
    ):
        sites.append((site.id, site.client.name, site.name))
    _style_sheet(sites, {"A": 12, "B": 32, "C": 30})
    _add_table(sites, "FacilityReferences")

    professions = workbook.create_sheet("Professions")
    professions.append(("profession_id", "profession_name"))
    for profession in profession_queryset.order_by("name", "id"):
        professions.append((profession.id, profession.name))
    _style_sheet(professions, {"A": 15, "B": 30})
    _add_table(professions, "FacilityProfessionReferences")
    return _workbook_bytes(workbook)


def _workbook_bytes(workbook):
    payload = BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _read_rows(uploaded, sheet_name, expected_headers):
    try:
        uploaded.seek(0)
        workbook = load_workbook(uploaded, read_only=True, data_only=True)
    except Exception as exc:
        raise RateWorkbookError(["The uploaded file is not a valid Excel workbook."]) from exc
    if sheet_name not in workbook.sheetnames:
        raise RateWorkbookError([f'The workbook must contain a "{sheet_name}" sheet.'])
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        actual_headers = tuple(
            str(value).strip() if value is not None else "" for value in next(rows)
        )
    except StopIteration as exc:
        raise RateWorkbookError([f'The "{sheet_name}" sheet is empty.']) from exc
    if actual_headers != expected_headers:
        raise RateWorkbookError([
            "The rate sheet columns were changed. Download a fresh workbook and keep "
            f"this exact header order: {', '.join(expected_headers)}."
        ])
    populated = []
    for row_number, row in enumerate(rows, start=2):
        values = tuple(row[: len(expected_headers)])
        if not any(value not in (None, "") for value in values):
            continue
        populated.append((row_number, values))
        if len(populated) > MAX_IMPORT_ROWS:
            raise RateWorkbookError([
                f"Import at most {MAX_IMPORT_ROWS} populated rate rows at a time."
            ])
    if not populated:
        raise RateWorkbookError(["Add at least one populated rate row."])
    return populated


def _integer(value, row_number, field_name, required=True):
    if value in (None, ""):
        if required:
            raise ValueError(f"Row {row_number}: {field_name} is required.")
        return None
    if isinstance(value, bool):
        raise ValueError(f"Row {row_number}: {field_name} must be a whole-number ID.")
    try:
        integer = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Row {row_number}: {field_name} must be a whole-number ID."
        ) from exc
    if str(value).strip() not in {str(integer), f"{integer}.0"} and not (
        isinstance(value, float) and value.is_integer()
    ):
        raise ValueError(f"Row {row_number}: {field_name} must be a whole-number ID.")
    if integer <= 0:
        raise ValueError(f"Row {row_number}: {field_name} must be greater than zero.")
    return integer


def _rate(value, row_number, field_name):
    try:
        decimal = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_number}: {field_name} must be a valid amount.") from exc
    if not decimal.is_finite():
        raise ValueError(f"Row {row_number}: {field_name} must be a valid amount.")
    if decimal < 0:
        raise ValueError(f"Row {row_number}: {field_name} must be zero or greater.")
    if decimal > MAX_RATE:
        raise ValueError(f"Row {row_number}: {field_name} is too large.")
    if decimal != decimal.quantize(CENT):
        raise ValueError(
            f"Row {row_number}: {field_name} can have at most two decimal places."
        )
    return decimal.quantize(CENT)


def import_client_rate_workbook(
    uploaded,
    rate_queryset,
    client_queryset,
    profession_queryset,
):
    rows = _read_rows(uploaded, "Client Rates", CLIENT_RATE_HEADERS)
    clients = {item.id: item for item in client_queryset}
    professions = {item.id: item for item in profession_queryset}
    existing_by_id = {item.id: item for item in rate_queryset}
    parsed, errors, seen = [], [], set()
    for row_number, row in rows:
        try:
            rate_id = _integer(row[0], row_number, "rate_id", required=False)
            client_id = _integer(row[1], row_number, "client_id")
            profession_id = _integer(row[3], row_number, "profession_id")
            pay_rate = _rate(row[5], row_number, "pay_rate")
            bill_rate = _rate(row[6], row_number, "bill_rate")
            client = clients.get(client_id)
            profession = professions.get(profession_id)
            if client is None:
                raise ValueError(
                    f"Row {row_number}: client_id {client_id} is unknown or outside your Department access."
                )
            if profession is None:
                raise ValueError(
                    f"Row {row_number}: profession_id {profession_id} is unknown."
                )
            key = (client_id, profession_id)
            if key in seen:
                raise ValueError(
                    f"Row {row_number}: this Client and role appears more than once."
                )
            seen.add(key)
            existing = existing_by_id.get(rate_id) if rate_id else None
            if rate_id and existing is None:
                raise ValueError(
                    f"Row {row_number}: rate_id {rate_id} is unknown or outside your Department access."
                )
            if existing and (
                existing.client_id != client_id
                or existing.profession_id != profession_id
            ):
                raise ValueError(
                    f"Row {row_number}: rate_id {rate_id} does not belong to this Client and role."
                )
            parsed.append((client, profession, pay_rate, bill_rate))
        except ValueError as exc:
            errors.append(str(exc))
    if errors:
        raise RateWorkbookError(errors)
    return _apply_client_rates(parsed)


@transaction.atomic
def _apply_client_rates(parsed):
    entries = []
    created_count = updated_count = 0
    for client, profession, pay_rate, bill_rate in parsed:
        rate, created = ClientProfessionRate.objects.select_for_update().get_or_create(
            client=client,
            profession=profession,
            defaults={
                "pay_rate": pay_rate,
                "bill_rate": bill_rate,
                "locally_managed": True,
            },
        )
        if created:
            created_count += 1
        else:
            rate.pay_rate = pay_rate
            rate.bill_rate = bill_rate
            rate.locally_managed = True
            try:
                rate.full_clean()
            except ValidationError as exc:
                raise RateWorkbookError(exc.messages) from exc
            rate.save(update_fields=[
                "pay_rate", "bill_rate", "locally_managed", "synced_at"
            ])
            updated_count += 1
        entries.append((rate, created))
    return RateImportSummary(created_count, updated_count, entries)


def import_site_rate_workbook(
    uploaded,
    rate_queryset,
    site_queryset,
    profession_queryset,
):
    rows = _read_rows(uploaded, "Facility Rates", SITE_RATE_HEADERS)
    sites = {item.id: item for item in site_queryset}
    professions = {item.id: item for item in profession_queryset}
    existing_by_id = {item.id: item for item in rate_queryset}
    parsed, errors, seen = [], [], set()
    for row_number, row in rows:
        try:
            rate_id = _integer(row[0], row_number, "rate_id", required=False)
            site_id = _integer(row[1], row_number, "site_id")
            profession_id = _integer(row[4], row_number, "profession_id")
            pay_rate = _rate(row[6], row_number, "pay_rate")
            bill_rate = _rate(row[7], row_number, "bill_rate")
            site = sites.get(site_id)
            profession = professions.get(profession_id)
            if site is None:
                raise ValueError(
                    f"Row {row_number}: site_id {site_id} is unknown or outside your Department access."
                )
            if profession is None:
                raise ValueError(
                    f"Row {row_number}: profession_id {profession_id} is unknown."
                )
            key = (site_id, profession_id)
            if key in seen:
                raise ValueError(
                    f"Row {row_number}: this facility and role appears more than once."
                )
            seen.add(key)
            existing = existing_by_id.get(rate_id) if rate_id else None
            if rate_id and existing is None:
                raise ValueError(
                    f"Row {row_number}: rate_id {rate_id} is unknown or outside your Department access."
                )
            if existing and (
                existing.site_id != site_id
                or existing.profession_id != profession_id
            ):
                raise ValueError(
                    f"Row {row_number}: rate_id {rate_id} does not belong to this facility and role."
                )
            parsed.append((site, profession, pay_rate, bill_rate))
        except ValueError as exc:
            errors.append(str(exc))
    if errors:
        raise RateWorkbookError(errors)
    return _apply_site_rates(parsed)


@transaction.atomic
def _apply_site_rates(parsed):
    entries = []
    created_count = updated_count = 0
    for site, profession, pay_rate, bill_rate in parsed:
        rate, created = SiteProfessionRate.objects.select_for_update().get_or_create(
            site=site,
            profession=profession,
            defaults={"pay_rate": pay_rate, "bill_rate": bill_rate},
        )
        if created:
            created_count += 1
        else:
            rate.pay_rate = pay_rate
            rate.bill_rate = bill_rate
            try:
                rate.full_clean()
            except ValidationError as exc:
                raise RateWorkbookError(exc.messages) from exc
            rate.save(update_fields=["pay_rate", "bill_rate", "updated_at"])
            updated_count += 1
        entries.append((rate, created))
    return RateImportSummary(created_count, updated_count, entries)
