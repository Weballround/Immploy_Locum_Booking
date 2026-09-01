from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client as DjangoClient, RequestFactory
from openpyxl import Workbook, load_workbook
from django.contrib import admin

from bookings.models import (
    Booking,
    Client,
    ClientProfessionRate,
    Invoice,
    InvoiceLine,
    Profession,
    Site,
    SiteProfessionRate,
    TimesheetLine,
)


CLIENT_HEADERS = [
    "rate_id",
    "client_id",
    "client_name",
    "profession_id",
    "profession_name",
    "pay_rate",
    "bill_rate",
]
SITE_HEADERS = [
    "rate_id",
    "site_id",
    "client_name",
    "site_name",
    "profession_id",
    "profession_name",
    "pay_rate",
    "bill_rate",
]


def workbook_upload(sheet_name, headers, rows, filename="rates.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    payload = BytesIO()
    workbook.save(payload)
    return SimpleUploadedFile(
        filename,
        payload.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.fixture
def rate_setup():
    client = Client.objects.create(name="Central Hospital", region="Gauteng")
    site = Site.objects.create(client=client, name="Main Ward")
    doctor = Profession.objects.create(name="Medical Officer")
    nurse = Profession.objects.create(name="Registered Nurse")
    client_rate = ClientProfessionRate.objects.create(
        client=client,
        profession=doctor,
        pay_rate="500.00",
        bill_rate="700.00",
    )
    site_rate = SiteProfessionRate.objects.create(
        site=site,
        profession=doctor,
        pay_rate="550.00",
        bill_rate="760.00",
    )
    return client, site, doctor, nurse, client_rate, site_rate


@pytest.fixture
def admin_browser():
    user = get_user_model().objects.create_superuser(
        username="rate-admin",
        password="local-test-password",
    )
    browser = DjangoClient()
    browser.force_login(user)
    return browser


def grant_rate_permission(user, codename, name):
    permission, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Booking),
        codename=codename,
        defaults={"name": name},
    )
    user.user_permissions.add(permission)


def limited_rate_admin(username, model, *rate_permissions):
    user = get_user_model().objects.create_user(username=username, is_staff=True)
    content_type = ContentType.objects.get_for_model(model)
    user.user_permissions.add(
        Permission.objects.get(
            content_type=content_type,
            codename=f"view_{model._meta.model_name}",
        ),
        Permission.objects.get(
            content_type=content_type,
            codename=f"change_{model._meta.model_name}",
        ),
    )
    permission_names = {
        "view_candidate_pay_rates": "Can view Candidate pay rates",
        "view_client_charge_rates": "Can view Client charges and profitability",
        "override_approved_rates": "Can override approved rates",
    }
    for codename in rate_permissions:
        grant_rate_permission(user, codename, permission_names[codename])
    return user


@pytest.mark.django_db
def test_rate_admin_fields_follow_independent_view_permissions():
    pay_user = limited_rate_admin(
        "pay-field-admin",
        ClientProfessionRate,
        "view_candidate_pay_rates",
    )
    charge_user = limited_rate_admin(
        "charge-field-admin",
        ClientProfessionRate,
        "view_client_charge_rates",
    )
    model_admin = admin.site._registry[ClientProfessionRate]
    factory = RequestFactory()

    pay_request = factory.get("/admin/bookings/clientprofessionrate/")
    pay_request.user = pay_user
    charge_request = factory.get("/admin/bookings/clientprofessionrate/")
    charge_request.user = charge_user

    assert "pay_rate" in model_admin.get_list_display(pay_request)
    assert "bill_rate" not in model_admin.get_list_display(pay_request)
    assert "bill_rate" in model_admin.get_list_display(charge_request)
    assert "pay_rate" not in model_admin.get_list_display(charge_request)


@pytest.mark.django_db
def test_default_model_change_permission_cannot_override_rates_without_explicit_capability():
    user = limited_rate_admin(
        "non-overriding-rate-admin",
        SiteProfessionRate,
        "view_candidate_pay_rates",
        "view_client_charge_rates",
    )
    request = RequestFactory().get("/admin/bookings/siteprofessionrate/")
    request.user = user
    model_admin = admin.site._registry[SiteProfessionRate]

    assert model_admin.has_view_permission(request)
    assert not model_admin.has_change_permission(request)
    assert not model_admin.has_rate_import_permission(request)


@pytest.mark.django_db
def test_pay_only_admin_cannot_download_workbook_containing_client_charges():
    user = limited_rate_admin(
        "pay-only-export-admin",
        ClientProfessionRate,
        "view_candidate_pay_rates",
    )
    browser = DjangoClient()
    browser.force_login(user)

    response = browser.get("/admin/bookings/clientprofessionrate/export-xlsx/")

    assert response.status_code == 403


@pytest.mark.django_db
def test_timesheet_line_admin_redacts_the_unpermitted_rate_field():
    user = limited_rate_admin(
        "pay-only-timesheet-line-admin",
        TimesheetLine,
        "view_candidate_pay_rates",
    )
    request = RequestFactory().get("/admin/bookings/timesheetline/")
    request.user = user
    model_admin = admin.site._registry[TimesheetLine]

    assert model_admin.has_view_permission(request)
    assert "pay_rate" in model_admin.get_list_display(request)
    assert "bill_rate" not in model_admin.get_list_display(request)


@pytest.mark.django_db
def test_invoice_admin_requires_client_charge_visibility_in_addition_to_model_permission():
    pay_only_user = limited_rate_admin(
        "pay-only-invoice-admin",
        Invoice,
        "view_candidate_pay_rates",
    )
    charge_user = limited_rate_admin(
        "charge-invoice-admin",
        Invoice,
        "view_client_charge_rates",
    )
    invoice_admin = admin.site._registry[Invoice]
    invoice_line_admin = admin.site._registry[InvoiceLine]
    factory = RequestFactory()

    pay_request = factory.get("/admin/bookings/invoice/")
    pay_request.user = pay_only_user
    charge_request = factory.get("/admin/bookings/invoice/")
    charge_request.user = charge_user

    assert not invoice_admin.has_view_permission(pay_request)
    assert invoice_admin.has_view_permission(charge_request)
    assert not invoice_line_admin.has_view_permission(pay_request)


@pytest.mark.django_db
def test_rate_admin_lists_excel_import_and_download_controls(admin_browser):
    client_rates = admin_browser.get("/admin/bookings/clientprofessionrate/")
    site_rates = admin_browser.get("/admin/bookings/siteprofessionrate/")

    assert client_rates.status_code == 200
    assert site_rates.status_code == 200
    assert b"Import Excel" in client_rates.content
    assert b"Download Excel" in client_rates.content
    assert b"Import Excel" in site_rates.content
    assert b"Download Excel" in site_rates.content


@pytest.mark.django_db
def test_client_rate_excel_export_contains_existing_rates_and_reference_sheets(
    admin_browser,
    rate_setup,
):
    client, _, doctor, _, client_rate, _ = rate_setup

    response = admin_browser.get(
        "/admin/bookings/clientprofessionrate/export-xlsx/"
    )

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Client Rates", "Clients", "Professions"]
    row = list(workbook["Client Rates"].iter_rows(min_row=2, values_only=True))[0]
    assert row == (
        client_rate.id,
        client.id,
        client.name,
        doctor.id,
        doctor.name,
        500,
        700,
    )


@pytest.mark.django_db
def test_client_rate_excel_import_updates_and_creates_rates_atomically(
    admin_browser,
    rate_setup,
):
    client, _, doctor, nurse, client_rate, _ = rate_setup
    upload = workbook_upload(
        "Client Rates",
        CLIENT_HEADERS,
        [
            [
                client_rate.id,
                client.id,
                client.name,
                doctor.id,
                doctor.name,
                525,
                735,
            ],
            [None, client.id, client.name, nurse.id, nurse.name, 420, 610],
        ],
    )

    response = admin_browser.post(
        "/admin/bookings/clientprofessionrate/import-xlsx/",
        {"file": upload},
        follow=True,
    )

    assert response.status_code == 200
    assert b"1 rate created and 1 rate updated" in response.content
    client_rate.refresh_from_db()
    assert str(client_rate.pay_rate) == "525.00"
    assert str(client_rate.bill_rate) == "735.00"
    created = ClientProfessionRate.objects.get(client=client, profession=nurse)
    assert str(created.pay_rate) == "420.00"
    assert str(created.bill_rate) == "610.00"


@pytest.mark.django_db
def test_invalid_client_rate_excel_rolls_back_every_row(
    admin_browser,
    rate_setup,
):
    client, _, doctor, nurse, client_rate, _ = rate_setup
    upload = workbook_upload(
        "Client Rates",
        CLIENT_HEADERS,
        [
            [
                client_rate.id,
                client.id,
                client.name,
                doctor.id,
                doctor.name,
                525,
                735,
            ],
            [None, client.id, client.name, nurse.id, nurse.name, -1, 610],
        ],
    )

    response = admin_browser.post(
        "/admin/bookings/clientprofessionrate/import-xlsx/",
        {"file": upload},
    )

    assert response.status_code == 200
    assert b"Row 3" in response.content
    assert b"pay_rate must be zero or greater" in response.content
    client_rate.refresh_from_db()
    assert str(client_rate.pay_rate) == "500.00"
    assert not ClientProfessionRate.objects.filter(
        client=client,
        profession=nurse,
    ).exists()


@pytest.mark.django_db
def test_site_rate_excel_import_updates_and_creates_facility_overrides(
    admin_browser,
    rate_setup,
):
    client, site, doctor, nurse, _, site_rate = rate_setup
    upload = workbook_upload(
        "Facility Rates",
        SITE_HEADERS,
        [
            [
                site_rate.id,
                site.id,
                client.name,
                site.name,
                doctor.id,
                doctor.name,
                575,
                790,
            ],
            [None, site.id, client.name, site.name, nurse.id, nurse.name, 440, 630],
        ],
    )

    response = admin_browser.post(
        "/admin/bookings/siteprofessionrate/import-xlsx/",
        {"file": upload},
        follow=True,
    )

    assert response.status_code == 200
    assert b"1 rate created and 1 rate updated" in response.content
    site_rate.refresh_from_db()
    assert str(site_rate.pay_rate) == "575.00"
    assert str(site_rate.bill_rate) == "790.00"
    created = SiteProfessionRate.objects.get(site=site, profession=nurse)
    assert str(created.pay_rate) == "440.00"
    assert str(created.bill_rate) == "630.00"


@pytest.mark.django_db
def test_rate_excel_import_requires_add_and_change_permissions(rate_setup):
    user = get_user_model().objects.create_user(
        username="read-only-rate-user",
        password="local-test-password",
        is_staff=True,
    )
    browser = DjangoClient()
    browser.force_login(user)
    upload = workbook_upload("Client Rates", CLIENT_HEADERS, [])

    response = browser.post(
        "/admin/bookings/clientprofessionrate/import-xlsx/",
        {"file": upload},
    )

    assert response.status_code == 403


@pytest.mark.django_db
def test_direct_admin_rate_edit_marks_client_rate_locally_managed(
    admin_browser,
    rate_setup,
):
    client, _, doctor, _, client_rate, _ = rate_setup

    response = admin_browser.post(
        f"/admin/bookings/clientprofessionrate/{client_rate.id}/change/",
        {
            "client": client.id,
            "profession": doctor.id,
            "pay_rate": "535.00",
            "bill_rate": "745.00",
            "_save": "Save",
        },
    )

    assert response.status_code == 302
    client_rate.refresh_from_db()
    assert client_rate.locally_managed is True


@pytest.mark.django_db
def test_rate_admin_models_are_registered_for_individual_editing():
    request = RequestFactory().get("/admin/bookings/clientprofessionrate/")
    request.user = get_user_model().objects.create_superuser(
        username="rate-field-admin",
        password="local-test-password",
    )
    assert ClientProfessionRate in admin.site._registry
    assert SiteProfessionRate in admin.site._registry
    assert "pay_rate" in admin.site._registry[ClientProfessionRate].get_fields(request)
    assert "bill_rate" in admin.site._registry[SiteProfessionRate].get_fields(request)
